import os
import operator
import binascii
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors


def Hex_to_String(list):
    Str_Result = ''
    for byte in list:
        byte = '%0#4x' % byte
        Str=byte.replace('0x','').upper()
        Str_Result=Str_Result+Str+' '
    return Str_Result

def Dir_Creat(path):
    #根据目录创建字典
    file_path=path
    file = open(path)
    Dir_list={}
    # print(file)
    for line in file.readlines():
        # print(line)
        if line.find('[')>=0:
            ID=line[line.find('[')+3:line.find('[')+7]
            # print(ID)
            Name=line[line.find('=')+2:].replace('\"','').replace(',','').replace('\n','')
            # print('Name=',Name)
            Dir_list[ID]=Name
    # print(Dir_list)
    return Dir_list





def Doip_Data_analyze(list):
    list_head=list[:4]
    list_target=[0x02,0xfd,0x80,0x01]
    if operator.eq(list_head,list_target):
        # print(list_head)
        # print('Diagnostic info=',list)
        return list
    else:
        # print('false')
        return None



def Doip_Packet_Data(list):

    length = list
    # print(length.__len__())
    Data = length[54:length.__len__()]
    # print('Packet Data=', Data)
    # print('Data(Hex)=',end='')
    # for i in Data:
        # print(hex(i), end=' ')
    return Data

def list_to_string(list):
    length =list
    result = ''


    # lastresult = 0
    for data in length:
        data='%0#4x' %data
        result = result + data+' '
    return result


def Byte_length(length_list):
    time = 6
    length = length_list
    lastresult = 0
    for size in length:
        # print(size)
        result = size * 16 ** time
        lastresult = result + lastresult
        time=time-2

    return lastresult
    # print(lastresult)



def Doip_Data_analyze(list):
    list_head=list[:4]
    list_target=[0x02,0xfd,0x80,0x01]
    if operator.eq(list_head,list_target):
        # print(list_head)
        # print('Diagnostic info=',list)
        Doip_version=list_to_string(list[:2])
        PL_Type=list_to_string(list[2:4])
        PL_length=Byte_length(list[4:8])
        Source_Address=Hex_to_String(list[8:10]).replace(' ','')
        Target_Address=Hex_to_String(list[10:12]).replace(' ','')
        UDS_Data=Hex_to_String(list[12:])

        result=[Doip_version,PL_Type,PL_length,Source_Address,Target_Address,UDS_Data]
        return result



def data2Asc(data):
    string =data
    res = ''
    for i in range(1, int(len(string)) + 1, 2):
        byte = string[i - 1:i + 1]
        char = chr(int(byte, 16))
        if char == ' ':
            char = '空'
        # print(byte)
        # print(char)
        res = res + char
    # print(res)

    return res+'('+str(int((i+1)/2))+'Byte)'


def add_Blank(string):
    i = 2
    new_string = ''
    for char in string:
        new_string = new_string + char
        i = i - 1
        if i == 0:
            new_string = new_string + ' '
            i = 2

    return new_string


def log_translate(line):
    line=line[len(line)-24:]
    line=line.replace(' ','')
    return line

def log_translate_DSA(line):
    line=line[len(line)-33:len(line)-9]
    line=line.replace(' ','')
    return line

def log_translate_DSA_Type2(line):
    line=line[len(line)-25:len(line)]
    line=line.replace(' ','')
    return line


def testlog_translate_DSA(line):
    location=line.find('d 8 ')
    line=line[location+3:location+28]
    line=line.replace(' ','')
    return line


def deal_with_line(line):
    log_line=line.replace(' ','')
    Frame_type = log_line[0:1]
    if Frame_type == '0':
        #     单帧
        datalen = int(log_line[1:2], 16)
        data = log_line[2:2 + datalen * 2]
        data.replace('\n', '')
        data = add_Blank(data)
        print_data ='        ' + 'len=' + str(datalen) + '         ' + data

        ECU_log.write(print_data + '\n')

    if Frame_type == '1':
        #     连续帧首帧
        datalen = int(log_line[1:4], 16)
        data_frame_num = int(datalen / 7) + 1
        data = log_line[4:4 + datalen * 2]
        data = data.replace('\n', '')

    if Frame_type == '2':
        #     连续帧后续
        data = data + log_line[2:]
        data = data.replace('\n', '')
        data_frame_num = data_frame_num - 1
        if data_frame_num == 1:
            data = add_Blank(data)
            if datalen > 9:
                if datalen < 100:
                    print_data = CANID_line + '        ' + 'len=' + str(datalen) + '        ' + data

            else:
                print_data = CANID_line + '        ' + 'len=' + str(datalen) + '         ' + data

            ECU_log.write(print_data + '\n')

    if Frame_type == '3':
        pass


def  Name_settle(filename,ori_addr,tar_addr):


    os.chdir(ori_addr)
    file=open(filename)

    os.chdir(tar_addr)
    new_file=open(filename,'w')

    for line in file.readlines():

        # new_line=line.replace(' ','')
        # new_line=new_line.replace('\n','')
        # new_line=new_line[len(new_line)-21:]
        # print(new_line)
        # new_file.write(new_line)
        # print(line)
        if line.find('\t')!=-1:
            if line != '\t\n':
                line=line.replace('\n','')
        if line!='\t\n':
            new_file.write(line)

    file.close
    new_file.close



def Copy_file(filename,ori_addr,tar_addr):
    os.chdir(ori_addr)
    file = open(filename)

    os.chdir(tar_addr)
    new_file=open(filename,'w')
    for line in file.readlines():
        new_file.write(line)
    file.close
    new_file.close



def ECU_log(filename,log_file):


    def add_Blank(string):
        i=2
        new_string=''
        for char in string:
            new_string=new_string+char
            i=i-1
            if i==0:
                new_string = new_string +' '
                i=2

        return new_string




    file=open(filename)


    for line in file.readlines():
        logfile = open(log_file)
        blank=line.index('\t')
        ECU_name=line[0:blank]
        ECU_name=ECU_name.replace('/','')+'.txt'
        ECU_log=open(ECU_name,'w')
        # print(ECU_name)
        reqID_num=line.index('0x')
        ECU_reqID=line[reqID_num+2:reqID_num+5]
        ECU_resID=line[reqID_num+7:reqID_num+10]
        # print(ECU_reqID)
        # print(ECU_resID)

        datalen=0
        data=''
        data_frame_num=0

        for log_line in logfile.readlines():

            log_line=log_line.replace(' ','')
            log_line = log_line[len(log_line) - 21:]
            CANID_line=log_line[0:4]
            CANData_line=log_line[4:]
            if CANID_line.find(ECU_reqID)>0 or CANID_line.find(ECU_resID)>0:
                Frame_type=log_line[4:5]
                if Frame_type=='0':
                    #     单帧
                    datalen=int(log_line[5:6],16)
                    data=log_line[6:6+datalen*2]
                    data.replace('\n','')
                    data = add_Blank(data)
                    print_data=CANID_line+'        '+'len='+str(datalen)+'         '+data

                    ECU_log.write(print_data + '\n')


                if Frame_type=='1':
                    #     连续帧首帧
                    datalen = int(log_line[5:8],16)
                    data_frame_num=int(datalen/7)+1
                    data = log_line[8:8 + datalen * 2]
                    data = data.replace('\n', '')




                if Frame_type=='2':
                #     连续帧后续
                    data=data+log_line[6:]
                    data=data.replace('\n', '')
                    data_frame_num=data_frame_num-1
                    if data_frame_num==1:
                        data = add_Blank(data)
                        if datalen>9:
                            if datalen<100:
                                print_data=CANID_line+'        '+'len='+str(datalen)+'        '+data

                        else:
                            print_data = CANID_line + '        ' + 'len=' + str(datalen) + '         ' + data


                        ECU_log.write(print_data + '\n')





                if Frame_type=='3':
                    pass







                # ECU_log.write(log_line)
        ECU_log.close
        logfile.close



class ECU_info():
    def __init__(self):
            self.name
            self.ReqID
            self.ResID

def ECU_list(namefile):
    namefile=open(namefile)
    for nameline in namefile.readlines():
        ECU_name=nameline[:nameline.index('\t')]
        ECU=ECU_info
        ECU.name=ECU_name
















def ECU_log_Male2(filename,log_file):


    def add_Blank(string):
        i=2
        new_string=''
        for char in string:
            new_string=new_string+char
            i=i-1
            if i==0:
                new_string = new_string +' '
                i=2

        return new_string




    file=open(filename)


    for line in file.readlines():
        logfile = open(log_file)
        blank=line.index('\t')
        ECU_name=line[0:blank]
        ECU_name=ECU_name.replace('/','')+'.txt'
        ECU_log=open(ECU_name,'w')
        # print(ECU_name)
        if line.find('0x')>-1:
            reqID_num=line.index('0x')
            ECU_reqID=line[reqID_num+2:reqID_num+5]
            ECU_resID=line[reqID_num+7:reqID_num+10]
        # print(ECU_reqID)
        # print(ECU_resID)

        datalen=0
        data=''
        data_frame_num=0

        for log_line in logfile.readlines():

            log_line=log_line.replace(' ','')
            # log_line = log_line[len(log_line) - 24:]
            # CANID_line=log_line[0:4]
            # CANData_line=log_line[len(log_line)-16:]
            RX_location=log_line.index('Rx')
            CANID_line=log_line[RX_location-3:RX_location]
            CANData_line=log_line[RX_location+4:]

            if CANID_line.find(ECU_reqID)>0 or CANID_line.find(ECU_resID)>0:
                Frame_type=CANData_line[0:1]
                if Frame_type=='0':
                    #     单帧
                    datalen=int(CANData_line[1:2],16)
                    data=CANData_line[2:2+datalen*2]
                    data.replace('\n','')
                    data = add_Blank(data)
                    print_data=CANID_line+'        '+'len='+str(datalen)+'         '+data

                    ECU_log.write(print_data + '\n')


                if Frame_type=='1':
                    #     连续帧首帧
                    datalen = int(CANData_line[1:4],16)
                    data_frame_num=int(datalen/7)+1
                    data = CANData_line[4:4 + datalen * 2]
                    data = data.replace('\n', '')




                if Frame_type=='2':
                #     连续帧后续
                    data=data+CANData_line[2:]
                    data=data.replace('\n', '')
                    data_frame_num=data_frame_num-1
                    if data_frame_num==1:
                        data = add_Blank(data)
                        if datalen>9:
                            if datalen<100:
                                print_data=CANID_line+'        '+'len='+str(datalen)+'        '+data

                        else:
                            print_data = CANID_line + '        ' + 'len=' + str(datalen) + '         ' + data


                        ECU_log.write(print_data + '\n')





                if Frame_type=='3':
                    pass







                # ECU_log.write(log_line)
        ECU_log.close
        logfile.close



def Analyze_Data(Data):
    Data=Data.upper()


    Data=Data.replace(' ','')
    # print(Data)
    SID=Data[0:2]
    if SID!='7F':
        statu='Right'
        if SID in('31','71','2F','6F'):
            SID_Type=Data[2:4]
            DID=Data[4:8]
            Data_data=Data[8:]
        elif SID in('10','50','11','51','27','67'):
            SID_Type=Data[2:4]
            DID=None
            Data_data=Data[4:]
        else:
            SID_Type=None
            DID=Data[2:6]
            Data_data=Data[6:]
        NRC=''

    else:
        # statu='Wrong'
        NRC=Data[4:6]
        if NRC=='78':
            statu = 'Right'
        else:
            statu='Wrong'
        DID=None
        SID_Type = None
        Data_data=None

    array = (statu, SID, DID,SID_Type, Data_data, NRC)


    return array





def log_result(log_name):


    ft = Font(color=colors.RED)
    wb = load_workbook(log_name)
    wb_target = load_workbook('result.xlsx')

    ws = wb.active
    ws_target = wb_target.active
    column_result = ws_target.max_column
    j = 2
    print(dir(ws))
    list = wb.sheetnames
    print(list)
    print(list.__len__())
    for name in list:
        ws = wb[name]
        # print(ws.title)
        ws_target.cell(j, 2).value = ws.title
        ws_target.cell(j + 1, 2).value = ws.title
        add_j = 1

        row = ws.max_row
        column = ws.max_column

        print(row, column)
        DID_in_list='false'
        for row_num in range(1, row + 1):
            DID_in_list = 'false'
            column_target = column_result

            DID = ws.cell(row_num, 4).value
            # print(DID)
            SID = ws.cell(row_num, 3).value
            Data = ws.cell(row_num, 6).value
            for search_ID in range(1, ws_target.max_column + 1):
                if DID == ws_target.cell(1, search_ID).value:
                    DID_in_list='true'
                    if SID == '62':
                        if DID in ['F187', 'F18A', 'F193', 'F195', 'F190', 'F112', 'F18C']:
                            try:
                                ws_target.cell(j + 1, search_ID).value = data2Asc(Data)
                            except:
                                ws_target.cell(j + 1, search_ID).value = Data
                                ws_target.cell(j + 1, search_ID).font = ft

                        else:
                            ws_target.cell(j + 1, search_ID).value = Data
                    if SID == '2E':
                        if DID in ['F187', 'F18A', 'F193', 'F195', 'F190', 'F112', 'F18C']:
                            try:
                                ws_target.cell(j + 1, search_ID).value = data2Asc(Data)
                            except:
                                ws_target.cell(j + 1, search_ID).value = Data
                                ws_target.cell(j + 1, search_ID).font = ft
                        else:
                            ws_target.cell(j, search_ID).value = Data

            if DID_in_list=='false':
                if SID in ['31','71','2E']:
                    # ws_target.cell(1, ws_target.max_column +1).value = DID
                    ws_target.cell(j,column_target+add_j).value=SID+DID
                    add_j=add_j+1
                    print('add DID==',DID)
                    DID_in_list='true'



        for search_ID in range(1, column_target + 1):
            if ws_target.cell(j + 1, search_ID).value != ws_target.cell(j, search_ID).value:
                if ws_target.cell(j, search_ID).value != None:
                    ws_target.cell(j + 1, search_ID).font = ft

        j = j + 2
    log_name=log_name.replace('.xlsx','_result.xlsx')
    print('name=',log_name)
    # ws_target_new=wb_target.create_sheet('list')
    ws_target_new=wb_target.copy_worksheet(ws_target)
    ws_target_new.title='New'
    for new_row in range(2,ws_target_new.max_row):
        for new_column in range(3,ws_target_new.max_column):
            if ws_target_new.cell(new_row,new_column).value!=None:
                ws_target_new.cell(new_row, new_column).value='*'


    wb_target.save(log_name)
    wb.close()
