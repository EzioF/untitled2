import os
import openpyxl
import test
path=os.getcwd()
print(path)
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

ft=Font(color=colors.RED)

filename='C:/Users/shenghuan.luo/PycharmProjects/untitled2/Project/Masterlist owner/EOL_AS-TREG-list_V1.5（涉及ECU清单）.xlsx'
Masterlist_filename='C:/Users/shenghuan.luo/PycharmProjects/untitled2/Project/Masterlist owner/Masterlist W4D3.xlsx'
wb=openpyxl.load_workbook(filename,data_only=True)
Masterlist_wb=openpyxl.load_workbook(Masterlist_filename,data_only=True)
# print(wb)
# print(Masterlist_wb)
list_ws=wb['2.0 ECU list']
# print(list_ws.title)

Masterlist_ws=Masterlist_wb['Masterlist W4D3']
# print(Masterlist_ws.title)

for list_ws_row in range(1,list_ws.max_row+1):
    Target_ECU=list_ws.cell(list_ws_row,2).value
    # print(Target_ECU)
    for Masterlist_ws_row in range(1,Masterlist_ws.max_row+1):
        Found_list_ECU=Masterlist_ws.cell(Masterlist_ws_row,3).value
        # print('Found_list_ECU',Found_list_ECU)
        if Target_ECU==Found_list_ECU:
            # print(Target_ECU)
            list_ws.cell(list_ws_row, 3, Masterlist_ws.cell(Masterlist_ws_row, 5).value)
            list_ws.cell(list_ws_row, 4, Masterlist_ws.cell(Masterlist_ws_row, 6).value)
            if list_ws.cell(list_ws_row,5).value!=Masterlist_ws.cell(Masterlist_ws_row,12).value:
                print('changed name=',list_ws.cell(list_ws_row,5).value)
                list_ws.cell(list_ws_row,5,Masterlist_ws.cell(Masterlist_ws_row,12).value)


                list_ws.cell(list_ws_row, 5).font=ft
            if list_ws.cell(list_ws_row, 6).value!=Masterlist_ws.cell(Masterlist_ws_row, 13).value:
                print('changed name=', list_ws.cell(list_ws_row, 6).value)
                list_ws.cell(list_ws_row,6, Masterlist_ws.cell(Masterlist_ws_row, 13).value)
                list_ws.cell(list_ws_row, 6).font=ft
            # list_ws.cell(list_ws_row, 11, Masterlist_ws.cell(Masterlist_ws_row, 29).value)
            print(Target_ECU)
            print('row',Masterlist_ws_row)
            print(Masterlist_ws.cell(Masterlist_ws_row,12).value)

Masterlist_wb.close()
wb.save('result.xlsx')

