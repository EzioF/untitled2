import os
import openpyxl
import test
path=os.getcwd()
new_path=path+'\log'
book={}

namefile=open('FE-4KB.txt')
# log_name='2018-12-12（综合电检设备）.log'
log_name=input('输入log文件路径=')
# log_filename=log_name
# log_excel_name=log_name+'.xlsx'

from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
#
# ft=Font(color=colors.RED)
#
wb=Workbook()
# i=1
for nameline in namefile.readlines():
    if nameline.find('0x')>-1:
        ECU_name=nameline[:nameline.index('\t')]
        Req_ID=nameline[nameline.index('0x')+2:nameline.index('0x')+5]
        Res_ID = nameline[nameline.index('0x') + 7:nameline.index('0x') + 10]
        print(ECU_name+Req_ID+Res_ID)
        log_file=open(log_name)
#         # log_file = open('log_VF11.txt')
#         log type: $07,$E4,$0C,$00,$00,$07,$EC
        KB_ID='$0'+Req_ID[0:1]+','+'$'+Req_ID[1:3]+',$0C,$00,$00,'+'$0'+Res_ID[0:1]+','+'$'+Res_ID[1:3]
        print('KB_ID=',KB_ID)
        book[KB_ID]=ECU_name


#
        sheet = wb.create_sheet()
#         # i=i+1
        sheet.title = ECU_name
#
# print(book.keys())
# print(book.values())

for line in log_file.readlines():
    for ID in book.keys():
        if line.find(ID)>=0:
            # print(line)
            sheet=wb[book[ID]]
    sheet.cell(sheet.max_row+1,1,line)




# for ws in wb:
#     if ws.cell(2,1).value==None:
#         wb.remove(ws)
wb.save('result.xlsx')
print('book=',book)
result_file=log_name.replace('.','')+'.xlsx'
wb.save(result_file)



