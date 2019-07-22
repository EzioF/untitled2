#_coding_=utf-8
import os
import test
import openpyxl

# print('test1')
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
#
ft=Font(color=colors.RED)
#
wb=Workbook()



ECU_list_patch='ECU address.txt'
ECU_list=test.Dir_Creat(ECU_list_patch)
# 创造ECU list

for ECU in ECU_list.keys():
    ECU_name=ECU_list[ECU]
    sheet = wb.create_sheet()
    #         # i=i+1
    sheet.title = ECU_name





Packet_length_list=[]
Packet_length=0
header=[]
packet=[]
header_flag=False

sheet_name=''
# result=test.Byte_length([00,00,00,74])
# print(result)
Packet_flag=False


# pcaps = rdpcap("YV1LC10D4J1388398_20181228_081404_DoIPData.pcap")

# print(help(scapy))
lineNumber=0
byteNumber=0
packetLeagth=0
next_PacketHeader=-1
next_Packet=-1
log_start_time=0.0
message_dif_time=0.0

# file=open('YV1LC10D4J1388398_20181228_081404_DoIPData.pcap','rb')
# file=open('0088111_20181205_112116_ECOS DoIP Log.pcap','rb')
log_path=input('输入log目录=')
file=open(log_path,'rb')



for line in  file.read():
    byteNumber=byteNumber+1
    # byte=ord(line)
    # print(hex(byte))
    # print (hex(line))
    if lineNumber==0:
        print('Pacp Header')
    # print('%0#4x' % (line,),end = ' ')

    if header_flag:
        header.append(line)

    if Packet_flag:
        packet.append(line)

    if lineNumber==23:
        # print()
        # print('Packet Header')
        header_flag=True
        Packet_flag=False






    if lineNumber==39:
        # print()
        header_flag=False
        Packet_flag = True
        # print('Packet header',header)
        Packet_length_list=header[8:12][::-1]
        Packet_timestamp=header[0:8]
        Packet_time_seconds=Packet_timestamp[0:4][::-1]
        Packet_time_microseconds = Packet_timestamp[4:8][::-1]

        log_start_time=test.Doip_time_stamp(Packet_time_seconds,Packet_time_microseconds)


        # print('Packetlist=',Packet_length_list)
        packetLeagth=test.Byte_length(Packet_length_list)
        # print('length=',packetLeagth)
        header=[]
        next_PacketHeader=lineNumber+packetLeagth

    if lineNumber==next_PacketHeader:
        # print()

        header_flag = True
        Packet_flag = False
        # print('Packet=',packet)
        # print('Packet Header')


        Packet_Data=test.Doip_Packet_Data(packet)
        Packet_Diagnostic_info=test.Doip_Data_analyze(Packet_Data)
        if Packet_Diagnostic_info!=None:
            # print("Diagnostic_info_hex=",Packet_Diagnostic_info)
            Source_Address=Packet_Diagnostic_info[3]
            Target_Address=Packet_Diagnostic_info[4]
            UDS_Data=Packet_Diagnostic_info[5]
            # print('Address=',Source_Address,Target_Address,UDS_Data)

            if Source_Address in ECU_list.keys() or Target_Address in ECU_list.keys():
                if Source_Address in ECU_list.keys():
                    sheet_name=ECU_list[Source_Address]
                if Target_Address in ECU_list.keys():
                    sheet_name=ECU_list[Target_Address]

                ws=wb[sheet_name]
                line_now=ws.max_row
                ws.cell(line_now+1,1,Source_Address)
                ws.cell(line_now+1, 2, Target_Address)
                # UDS data 12
                ws.cell(line_now+1, 12, UDS_Data)

                # UDS长度
                ws.cell(line_now + 1, 11).value=str(int(len(UDS_Data.replace(' ',''))/2))+'Bytes'

                # time

                ws.cell(line_now+1,8,message_dif_time)
                # test.Analyze_Data(UDS_Data)
        # UDS数据解析
                result = test.Analyze_Data(UDS_Data)
                # result = (statu, SID, DID,SID_Type, Data_data, NRC)
                if result[0] == 'Wrong':
                    # sheet.cell(j, 10, result[0]).font = ft
                    ws.cell(line_now + 1, 10, result[0]).font=ft
                else:
                    # sheet.cell(j, 10, result[0])
                    ws.cell(line_now + 1, 10, result[0])
                # sheet.cell(j, 3, result[1])
                ws.cell(line_now + 1, 3, result[1])
                # sheet.cell(j, 4, result[2])
                ws.cell(line_now + 1, 4, result[2])
                # sheet.cell(j, 5, result[3])
                ws.cell(line_now + 1, 5, result[3])
                # sheet.cell(j, 6, result[4])
                ws.cell(line_now + 1, 6, result[4])




        # print()



        # print('Packet length=',packet.__len__())
        # print('lineNumber=',lineNumber)
        packet=[]

        next_Packet=lineNumber+16

    if lineNumber==next_Packet:
        # print()
        header_flag = False
        Packet_flag = True
        # print('Packet header', header)
        Packet_length_list = header[8:12][::-1]
        # 解析时间戳
        Packet_timestamp=header[0:8]
        Packet_time_seconds=Packet_timestamp[0:4][::-1]
        Packet_time_microseconds = Packet_timestamp[4:8][::-1]
        # print('timestamp',Packet_time_seconds,Packet_time_microseconds)
        message_real_time=test.Doip_time_stamp(Packet_time_seconds,Packet_time_microseconds)
        message_dif_time=message_real_time-log_start_time



        # print('Packetlist=', Packet_length_list)
        packetLeagth = test.Byte_length(Packet_length_list)
        # print('length=', packetLeagth)
        header = []
        next_PacketHeader = lineNumber+ packetLeagth
    #
    # if lineNumber>1500:
    #     break




    lineNumber=lineNumber+1
    # if lineNumber>15:
    #     print()
    #     lineNumber=0
# print()
# print("ByteNumber=",byteNumber)
# print("header",header)
# print(ECU_list)

for ws in wb:
    if ws.cell(4,1).value==None:
        wb.remove(ws)

# wb.save('result.xlsx')
#
result_file=log_path+'.xlsx'
wb.save(result_file)
wb.close()
