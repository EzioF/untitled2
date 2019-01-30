import os
import openpyxl
import test
path=os.getcwd()
print(path)
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

ft=Font(color=colors.RED)

filename='C:\\Users\\shenghuan.luo\\PycharmProjects\\untitled2\\Project\\SW changed name\\SW-TREG State.xlsx'

wb=openpyxl.load_workbook(filename)
last_writer='old'
print(wb)
flag=0
list_ws=wb.active
print(list_ws.title)
Target_name=''
# Masterlist_ws=Masterlist_wb['ECU Details']
# print(Masterlist_ws.title)
#
for list_ws_row in range(1,list_ws.max_row+1):
    line_name=list_ws.cell(list_ws_row,1).value
    writer=list_ws.cell(list_ws_row,2).value
    ori_write=['shenweijian','xialiping']
    if line_name.find('Parameter')>=0 and line_name.find('Test')<0:
        Target_name=line_name
        last_writer = 'old'
        flag=1
    if writer in ori_write:
        continue
    if flag:
        print('ECU',Target_name)
        flag=0

    if last_writer !=writer:

        print(writer)
    last_writer=writer