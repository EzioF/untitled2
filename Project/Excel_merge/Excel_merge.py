import openpyxl
import os

path=input("路径=")

list=os.listdir(path)

excle_list=[]

# print(os.listdir(path))
for name in list:
    if name.find('xlsx')>=0:
        excle_list.append(name)

# print(excle_list)

from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

ft=Font(color=colors.RED)

Result_Workbook=Workbook()
for excle in excle_list:
    excle_path=path+'\\'+excle
    # print(excle_path)
    excel_wb=openpyxl.load_workbook(excle_path)
    excel_wb_name=excel_wb.sheetnames
    for sheet_name in excel_wb_name:
        if sheet_name in Result_Workbook.sheetnames:
#             获得所有的Excel到目标Excel
            pass
        else:
            sheet=Result_Workbook.create_sheet()
            sheet.title=sheet_name
        Result_Worksheet=Result_Workbook[sheet_name]

        # print('class',Result_Worksheet.type)

        # print('sheet_tilt',Result_Worksheet.title)
        for excel_row in range(1,excel_wb[sheet_name].max_row):
            Row=Result_Worksheet.max_row+1
            # print("Row=",Row)
            Result_Worksheet.cell(Row,1).value=excel_wb[sheet_name].cell(excel_row,1).value
            Result_Worksheet.cell(Row, 2).value = excel_wb[sheet_name].cell(excel_row, 2).value
            Result_Worksheet.cell(Row, 3).value = excel_wb[sheet_name].cell(excel_row, 3).value

Result_Path=path+'\\result.xlsx'
Result_Workbook.save(Result_Path)



