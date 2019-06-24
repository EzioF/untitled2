import openpyxl
import test

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill



def EOL_analysis(test_item,work_sheet):
    Result = None
    if test_item=='Information Check Type1':
        SID='22'
        SID_Type=None
        DID=['F187','F18A','F193','F195']
        for row in range(1,work_sheet.max_row+1):
            Log_Data=work_sheet.cell(row,12).value
            #  array = (statu, SID, DID,SID_Type, Data_data, NRC)
            log_result=test.Analyze_Data(Log_Data)
            print('log_result=',log_result)

            if log_result[1]==SID and log_result[2] in DID:
                Result= 'F187 YES'
                break
            else:
                Result = None

    if test_item=='F110 Write Type1':
        SID = '2E'
        SID_Type = None
        DID = 'F110'
        for row in range(1, work_sheet.max_row + 1):
            Log_Data = work_sheet.cell(row, 12).value
            #  array = (statu, SID, DID,SID_Type, Data_data, NRC)
            log_result = test.Analyze_Data(Log_Data)
            print('log_result=', log_result)

            if log_result[1] == SID and log_result[2] in DID:
                Result= 'F110 YES'
                break
            else:
                 Result= None




    return Result





log_sheet_path='000127.xlsx'
result_log_WB=load_workbook(log_sheet_path)
log_sheet_name=result_log_WB.sheetnames

# EOL list

EOL_list_sheet_path='GEEA1.0 EOL List.xlsx'
list_result_WB=load_workbook(EOL_list_sheet_path)
list_result_WS=list_result_WB.active






for ECU_sheet in log_sheet_name:
    # print(ECU_sheet)
    if ECU_sheet!='Function':

        for list_row in range(1,list_result_WS.max_row):
            if ECU_sheet==list_result_WS.cell(list_row,1).value:
                for ECU_row in range(0,100):
                    if list_result_WS.cell(list_row+ECU_row,1).value!=ECU_sheet:
                        break
                    else:
                        # 获取测试名称
                        test_item=list_result_WS.cell(list_row+ECU_row,3).value
                        # 分析测试
                        list_result_WS.cell(list_row + ECU_row, 4).value=EOL_analysis(test_item,result_log_WB[ECU_sheet])


list_result_WB.save('log_result.xlsx')





