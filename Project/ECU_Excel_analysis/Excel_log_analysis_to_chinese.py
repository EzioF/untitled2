import openpyxl
import test

# 配置
SID_Column=3
DID_Column=4
Sub_ID_Column=5
Data_Column=6
Chinese_Column=14

from openpyxl import load_workbook
from openpyxl import Workbook

def UDS_to_Chinese(SID,DID,Sub_ID,Data):
    Sub_ID_Trans=''
    SID_Trans=''
    DID_Trans=''
    Data_Trans=''
    # 会话模式切换
    if SID=='10':
        SID_Trans='会话模式进入'
        if Sub_ID=='01':
            Sub_ID_Trans='默认'
        if Sub_ID == '03':
            Sub_ID_Trans='拓展'
    if SID=='50':
        SID_Trans = '会话模式进入成功'
        if Sub_ID == '01':
            Sub_ID_Trans = '默认'
        if Sub_ID == '03':
            Sub_ID_Trans = '拓展'

    # 读写
    if SID=='22':
        SID_Trans='读取'

        if DID=='F187':
            DID_Trans='零件号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F18A':
            DID_Trans='供应商代码'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F193':
            DID_Trans='硬件版本号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F195':
            DID_Trans='软件版本号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F101':
            DID_Trans='功能配置'
            if Data != None:
                Data_Trans = Data
        if DID=='F110':
            DID_Trans='网络配置'
            if Data != None:
                Data_Trans = Data
        if DID=='F18C':
            DID_Trans='序列号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)

    if SID=='62':
        SID_Trans='读取成功'

        if DID=='F187':
            DID_Trans='零件号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F18A':
            DID_Trans='供应商代码'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F193':
            DID_Trans='硬件版本号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F195':
            DID_Trans='软件版本号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)
        if DID=='F101':
            DID_Trans='功能配置'
            if Data != None:
                Data_Trans = Data
        if DID=='F110':
            DID_Trans='网络配置'
            if Data != None:
                Data_Trans = Data
        if DID=='F18C':
            DID_Trans='序列号'
            if Data != None:
                Data_Trans = test.data2Asc(Data)





    Chinese=Sub_ID_Trans+SID_Trans+DID_Trans+': '+Data_Trans

    return Chinese
# UDS翻译结束


log_excel_WB=load_workbook('000127.xlsx')
# print(log_excel_WB.sheetnames)
log_sheet_name_list=log_excel_WB.sheetnames
print( log_sheet_name_list)

result_WB=Workbook()
result_WS=result_WB.create_sheet('New')

# 遍历sheet页

log_work_sheet=log_excel_WB[log_sheet_name_list[2]]
print(log_work_sheet.title)

for log_row in range(1,log_work_sheet.max_row+1):
    # print(log_row)
    SID=log_work_sheet.cell(log_row,SID_Column).value
    DID = log_work_sheet.cell(log_row, DID_Column).value
    Sub_ID = log_work_sheet.cell(log_row, Sub_ID_Column).value
    Data = log_work_sheet.cell(log_row, Data_Column).value
    print(UDS_to_Chinese(SID,DID,Sub_ID,Data))

    log_work_sheet.cell(log_row,Chinese_Column).value=UDS_to_Chinese(SID,DID,Sub_ID,Data)

for row in range(1,log_work_sheet.max_row+1):
    for column in range(1,log_work_sheet.max_column+1):
        result_WS.cell(row,column).value=log_work_sheet.cell(row,column).value



result_WB.save('Result.xlsx')



