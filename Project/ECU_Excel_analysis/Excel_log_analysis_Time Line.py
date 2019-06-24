import openpyxl
# 变量
ECU_time_column=1
Result_Row=2

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill

fill=PatternFill('solid',fgColor='2E8B57')

# sheet analysis
# 时间转换
def string_time_To_second(time):
    if time.find(':')>=0:
        minute=time[:time.index(':')]
        rest_string=time[time.index(':'):]
        second=rest_string[1:3]
        # print('sec=',second)
        sec_string=str(int(minute)*60+int(second))

        return sec_string
    else:
        return None



def log_analysis(sheet,result_sheet):
    # 行命名
    result_sheet.cell(Result_Row, 1).value = sheet.title


    for row in range(1,sheet.max_row+1):
#         确认行
        #ECU log分析,获取时间
        time=sheet.cell(row,8).value

        Sec_time=string_time_To_second(time)
        # print('sec-time',Sec_time)
        for time_column in range(2,int(result_sheet.max_column)+1):
            time_axis=result_sheet.cell(1,time_column).value
            # print('time_axis',time_axis)
            if time_axis!=None:
                if int(Sec_time)<=int(time_axis):
                    ECU_time_column=time_column
                    break
                else:
                    pass
            else:
                pass
#             涂颜色
        result_sheet.cell(Result_Row,ECU_time_column).fill=fill











# sheet analysis end

sheet_path=input('输入Excel路径')

log_Worksheet=load_workbook(sheet_path)
log_sheet_name=log_Worksheet.sheetnames

result_worksheet=Workbook()

result_sheet=result_worksheet.active
print(result_sheet.title)
# 创造时间轴
for time in range(2,1000,2):
    # print('time=',time)
    result_sheet.cell(1,time/2+1).value=time
result_sheet.cell(1,1).value='ECU'
# 时间轴结束

# open sheet
for sheetname in log_sheet_name:
    sheet=log_Worksheet[sheetname]
    print(sheet.title)

    log_analysis(sheet,result_sheet)

    Result_Row=Result_Row+1




# save
result_sheet_path=sheet_path.replace('.xlsx','Time Line.xlsx')
result_worksheet.save(result_sheet_path)
log_Worksheet.close()
