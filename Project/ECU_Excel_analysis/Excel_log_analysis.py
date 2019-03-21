import openpyxl

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
# sheet analysis
# 时间转换
def string_time_To_second(time):
    if time.find(':')>=0:
        minute=time[:time.index(':')]
        rest_string=time[time.index(':'):]
        second=rest_string[1:3]
        # print('sec=',second)
        sec_string=str(int(minute)*60+int(second))

        return sec_string
    else:
        return None



def log_analysis(sheet,result_sheet):
    for row in range(1,sheet.max_row+1):
#         确认行
        #ECU log分析,获取时间
        time=sheet.cell(row,8).value

        Sec_time=string_time_To_second(time)
        print('sec-time',Sec_time)
        # for time_column in range(1,result_sheet.max_column+1):
        #     print('sec=',)







# sheet analysis end

log_Worksheet=load_workbook('000127.xlsx')
log_sheet_name=log_Worksheet.sheetnames

result_worksheet=Workbook()

result_sheet=result_worksheet.active
print(result_sheet.title)
# 创造时间轴
for time in range(2,200,2):
    # print('time=',time)
    result_sheet.cell(1,time/2+1).value=time
result_sheet.cell(1,1).value='ECU'
# 时间轴结束

# open sheet
for sheetname in log_sheet_name:
    sheet=log_Worksheet[sheetname]
    print(sheet.title)

    log_analysis(sheet,result_sheet)




# save
result_worksheet.save('result.xlsx')
log_Worksheet.close()
