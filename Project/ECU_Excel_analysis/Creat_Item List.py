import openpyxl

row=1
from openpyxl import Workbook
# 获取ECU信息
ECU_name_file=open('name_ECU.txt')

class ECU_info:
    def __init__(self,ECU_name,Req_ID,Res_ID):
        self.Name=ECU_name
        self.ReqID=Req_ID
        self.ResID=Res_ID

namelist=[]



for nameline in ECU_name_file.readlines():
    if nameline.find('\"')>=0:
        ECU=nameline[:nameline.index('\t')]
        Req_ID = nameline[nameline.index('0x') + 2:nameline.index('0x') + 5]
        Res_ID = nameline[nameline.index('0x') + 7:nameline.index('0x') + 10]
        ECU_infomation=ECU_info(ECU,Req_ID,Res_ID)
        namelist.append(ECU_infomation)

print(namelist)
# 获取ECU信息

ECU_list_WB=Workbook()

ECU_list_WS=ECU_list_WB.active

print(ECU_list_WS.title)

# 创建平台化清单
for ECU_list in namelist:
#     ECU增加测试内容
    #ECU名字
    ECU_list_WS.cell(row,1).value=ECU_list.Name
    ECU_list_WS.cell(row+1,1).value=ECU_list.Name
    ECU_list_WS.cell(row+2, 1).value = ECU_list.Name
    ECU_list_WS.cell(row+3, 1).value = ECU_list.Name
    ECU_list_WS.cell(row+4, 1).value = ECU_list.Name
    ECU_list_WS.cell(row+5, 1).value = ECU_list.Name
    ECU_list_WS.cell(row+6, 1).value = ECU_list.Name

    # 地址
    ECU_list_WS.cell(row,2).value='0x'+ECU_list.ReqID+'0x'+ECU_list.ResID
    # 测试内容
        # 件号检查
    ECU_list_WS.cell(row,3).value='Information Check Type1'
        #车辆信息写入
    ECU_list_WS.cell(row + 1,3).value='Information Write Type1'
        #配置F110
    ECU_list_WS.cell(row + 2, 3).value = 'F110 Write Type1'
        #配置F101
    ECU_list_WS.cell(row + 3, 3).value = 'F101 Write Type1'
        #备份配置
    ECU_list_WS.cell(row + 4, 3).value = 'Buckup Config Write Type1'
        #DTC清读

    ECU_list_WS.cell(row + 5, 3).value = 'DTC Clear'
    ECU_list_WS.cell(row + 6, 3).value = 'DTC Read'




    # 下一个ECU
    row=row+7

ECU_list_WB.save('GEEA1.0 EOL List.xlsx')








